#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Extract W-shape data from AISC Database Excel file and generate CSV
"""

from openpyxl import load_workbook
import csv
import os

def extract_w_shapes(excel_file, csv_file):
    """
    Extract W-shape data from AISC database Excel file

    Parameters:
    - excel_file: Path to AISC Shapes Database Excel file
    - csv_file: Path to output CSV file
    """
    print(f"Reading Excel file: {excel_file}")
    wb = load_workbook(excel_file, data_only=True)
    sheet = wb['Database v16.0']

    # Find column indices for required properties
    header = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
    col_indices = {}
    required_cols = {
        'Type': 'Type',
        'AISC_Manual_Label': 'designation',
        'd': 'd',
        'bf': 'bf',
        'tw': 'tw',
        'tf': 'tf',
        'Zx': 'Zx',
    }

    # Find column indices
    for excel_col, csv_col in required_cols.items():
        try:
            col_indices[csv_col] = header.index(excel_col)
        except ValueError:
            print(f"Warning: Column '{excel_col}' not found in header")

    print(f"Column indices: {col_indices}")

    # Extract W-shape data
    w_shapes = []
    total_rows = sheet.max_row

    print(f"Processing {total_rows} rows...")

    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        # Check if this is a W-shape
        shape_type = row[col_indices['Type']] if col_indices.get('Type') is not None else None

        if shape_type != 'W':
            continue

        # Extract required properties
        try:
            designation = row[col_indices['designation']] or ""
            d = float(row[col_indices['d']]) if row[col_indices['d']] is not None else 0
            bf = float(row[col_indices['bf']]) if row[col_indices['bf']] is not None else 0
            tw = float(row[col_indices['tw']]) if row[col_indices['tw']] is not None else 0
            tf = float(row[col_indices['tf']]) if row[col_indices['tf']] is not None else 0
            zx = float(row[col_indices['Zx']]) if row[col_indices['Zx']] is not None else 0

            # Only include valid entries
            if designation and d > 0 and bf > 0 and tw > 0 and tf > 0 and zx > 0:
                w_shapes.append({
                    'designation': designation,
                    'd': d,
                    'bf': bf,
                    'tw': tw,
                    'tf': tf,
                    'Zx': zx,
                })

                if len(w_shapes) % 100 == 0:
                    print(f"  Processed {len(w_shapes)} W-shapes so far...")

        except (ValueError, IndexError) as e:
            # Skip invalid rows
            continue

    print(f"Found {len(w_shapes)} W-shapes")

    # Write to CSV
    print(f"Writing to CSV: {csv_file}")
    with open(csv_file, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=['designation', 'd', 'bf', 'tw', 'tf', 'Zx'])
        writer.writeheader()
        writer.writerows(w_shapes)

    print(f"Successfully wrote {len(w_shapes)} W-shapes to {csv_file}")

    # Print some statistics
    if w_shapes:
        print("\nShape statistics:")
        print(f"  Total: {len(w_shapes)}")
        print(f"  Depth range: {min(s['d'] for s in w_shapes):.1f} to {max(s['d'] for s in w_shapes):.1f} in")
        print(f"  Weight range: {min(float(s['designation'].split('X')[1]) for s in w_shapes):.0f} to {max(float(s['designation'].split('X')[1]) for s in w_shapes):.0f} plf")

        # Show first and last few entries
        print("\nFirst 5 entries:")
        for s in w_shapes[:5]:
            print(f"  {s['designation']}: d={s['d']:.2f}, bf={s['bf']:.2f}, tw={s['tw']:.3f}, tf={s['tf']:.3f}, Zx={s['Zx']:.1f}")

        print("\nLast 5 entries:")
        for s in w_shapes[-5:]:
            print(f"  {s['designation']}: d={s['d']:.2f}, bf={s['bf']:.2f}, tw={s['tw']:.3f}, tf={s['tf']:.3f}, Zx={s['Zx']:.1f}")


if __name__ == "__main__":
    # File paths
    script_dir = os.path.dirname(os.path.abspath(__file__))
    excel_file = os.path.join(script_dir, "aisc-shapes-database-v160-2.xlsx")
    csv_file = os.path.join(script_dir, "aisc_w_shapes.csv")

    # Extract data
    extract_w_shapes(excel_file, csv_file)
    print("\nDone!")
